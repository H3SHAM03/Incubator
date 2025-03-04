from PyQt5 import QtWidgets, uic, QtCore, QtGui
from pyqtgraph import PlotWidget, plot
from PyQt5.QtWidgets import QVBoxLayout,QMessageBox
import pandas as pd
import sys
import os
import res1
import openpyxl
from datetime import datetime
import serial
import random
import threading
import time
import pygame
from datetime import datetime

class Stopwatch():
    def __init__(self):
        self.StartTime = 0
        self.EndTime = 0
        self.TimePassed = 0

    def start(self):
        self.StartTime = time.time()

    def secondsPassed(self):
        self.EndTime = time.time()
        self.TimePassed = self.EndTime - self.StartTime
        return self.TimePassed
    
    def reset(self):
        self.StartTime = 0
        self.EndTime = 0
        self.TimePassed = 0

class UserInfo():
	def __init__(self,username= None):
		self.username = username

 
	def run(self):
		self.isRunning = True
		if self._target is not None:
			self._return = self._target(*self._args, **self._kwargs)
		self.isRunning = False
		self.returnValue = self._return

session = UserInfo()
mode = 0

class LoginWindow(QtWidgets.QMainWindow):
	def __init__(self):
		super().__init__()
		self.ui = uic.loadUi("login.ui", self)
		self.setWindowFlags(QtCore.Qt.FramelessWindowHint)
		self.setAttribute(QtCore.Qt.WA_TranslucentBackground)
		self.ui.stackedWidget.setCurrentIndex(0)
		self.isLogin = True

		self.ui.pushButton_12.clicked.connect(self.minimize)
		self.ui.pushButton_10.clicked.connect(self.minimize)
		self.ui.pushButton_9.clicked.connect(self.exit)
		self.ui.pushButton_14.clicked.connect(self.exit)
		self.ui.pushButton_2.clicked.connect(self.SwitchLogin)
		self.ui.pushButton_6.clicked.connect(self.SwitchLogin)
		self.ui.pushButton.clicked.connect(self.Login)
		self.ui.pushButton_5.clicked.connect(self.Register)
		pygame.mixer.init()

	def SwitchLogin(self):
		self.isLogin = not self.isLogin
		if self.isLogin:
			self.ui.stackedWidget.setCurrentIndex(0)
		else:
			self.ui.stackedWidget.setCurrentIndex(1)

	def Login(self):
		username = self.ui.lineEdit.text()
		password = self.ui.lineEdit_2.text()
		users = pd.read_excel("assets\\Database\\users.xlsx", sheet_name=['Babies'])
		babies = users['Babies']
		found = False
		real_pw = ''
		for i,p in zip(babies['username'],babies['password']):
			if username == i:
				found = True
				real_pw = p
				break

		if found and (real_pw == password):
			self.ui.label_17.setText("Welcome!")
			global mode,session
			session = UserInfo(username=username)
			win = MainWindow()
			win.setWindowTitle("Incubator System")
			win.setWindowIcon(QtGui.QIcon("assets\\865969.png"))
			win.show()
			self.close()
		elif found and (real_pw != password):
			self.ui.label_17.setText("Wrong Password, Please Try Again.")
		elif not found:
			self.ui.label_17.setText("Username doesn't exist.")


	def Register(self):
		username = self.ui.lineEdit_5.text()
		password = self.ui.lineEdit_6.text()
		confirm = self.ui.lineEdit_9.text()
		users = pd.read_excel("assets\\Database\\users.xlsx", sheet_name=['Babies'])
		babies = users['Babies']
		found = False
		for i in babies['username']:
			if username == i:
				found = True
				break
		if ' ' in username:
			self.ui.label_18.setText("Username can't contain spaces.")
		elif username == '' or password == '' or confirm == '':
			self.ui.label_18.setText("Please fill all requirements.")
		elif password != confirm:
			self.ui.label_18.setText("Passwords don't match.")
		elif found:
			self.ui.label_18.setText("Username is taken.")
		else:
			book = openpyxl.load_workbook('assets\\Database\\users.xlsx')
			babies = book["Babies"]
			babies.append([username,password,'100','0','100','0','400','0'])
			book.save(os.getcwd() + "\\assets\\Database\\users.xlsx")
			self.ui.label_18.setText("Registered Successfully.")
		
	def exit(self):
		self.close()

	def minimize(self):
		self.showNormal()
		self.showMinimized()


class MainWindow(QtWidgets.QMainWindow):
	def __init__(self):
		# modes: Admin , Patient , Nurse
		super().__init__()
		self.ui = uic.loadUi("GUI.ui", self)
		self.ui.stackedWidget_2.setCurrentIndex(0)
		self.sw = Stopwatch()
		users = pd.read_excel("assets\\Database\\users.xlsx", sheet_name=['Babies'])
		babies = users['Babies']
		for i,baby in enumerate(babies['username']):
			if baby == session.username:
				self.ui.label_18.setText(str(babies['minTemp'][i]))
				self.ui.label_19.setText(str(babies['maxTemp'][i]))
				self.ui.label_20.setText(str(babies['minHum'][i]))
				self.ui.label_21.setText(str(babies['maxHum'][i]))
				self.ui.label_23.setText(str(babies['minBPM'][i]))
				self.ui.label_25.setText(str(babies['maxBPM'][i]))
				break
		self.TempGraph = PlotWidget()
		self.HumGraph = PlotWidget()
		self.BPMGraph = PlotWidget()
		self.Tempx = []
		self.Tempy = []
		self.Humx = []
		self.Humy = []
		self.BPMx = []
		self.BPMy = []
		layout1=QVBoxLayout()
		layout1.addWidget(self.TempGraph)
		layout2=QVBoxLayout()
		layout2.addWidget(self.HumGraph)
		layout3=QVBoxLayout()
		layout3.addWidget(self.BPMGraph)
		self.ui.widget_2.setLayout(layout3)
		self.ui.widget_4.setLayout(layout1)
		self.ui.widget_5.setLayout(layout2)
		self.TempGraph.addLegend()
		self.HumGraph.addLegend()
		self.BPMGraph.addLegend()
		self.Tempdata_line = self.TempGraph.plot(pen='r',name='Temperature')
		self.Humdata_line = self.HumGraph.plot(pen='b',name='Humidity')
		self.BPMdata_line = self.BPMGraph.plot(pen='g',name='BPM')
		self.TempGraph.setYRange(0,50)
		self.HumGraph.setYRange(0,100)
		self.BPMGraph.setYRange(0,200)
		self.TempGraph.setXRange(0,10)
		self.HumGraph.setXRange(0,10)
		self.BPMGraph.setXRange(0,10)
		try:
			self.serialPort = serial.Serial(port='COM8', baudrate=9600)
		except:
			self.connected = False
		else:
			self.connected = True
		self.timepassed = 0
		if self.connected:
			self.ReadThread = threading.Thread(target= self.Reading)
			self.ReadThread.start()
		else:
			self.ui.label_3.setText("Bluetooth Connection Failed.")

		self.ui.exit.clicked.connect(self.LogOut)
		self.ui.exit_2.clicked.connect(self.LogOut)
		self.ui.Monitoring.clicked.connect(self.MonitoringMode)
		self.ui.Statistics.clicked.connect(self.StatisticsMode)
		self.ui.Monitoring2.clicked.connect(self.MonitoringMode)
		self.ui.Statistics2.clicked.connect(self.StatisticsMode)
		self.ui.Jaundice.stateChanged.connect(self.DetectJaundice)

		self.ui.full_menu_2.setVisible(False)

	def Reading(self):
		self.sw.start()
		while self.connected:
			data = self.serialPort.readline().decode().strip()
			if  data and data != '':
				print(data)
				data = data.split('/')
				self.ReadThread.returnValue = data
				self.Tempx.append(self.sw.secondsPassed())
				self.Tempy.append(float(data[0]))
				self.Tempdata_line.setData(self.Tempx,self.Tempy)
				self.Humx.append(self.sw.secondsPassed())
				self.Humy.append(int(data[1]))
				self.Humdata_line.setData(self.Humx,self.Humy)
				self.BPMx.append(self.sw.secondsPassed())
				self.BPMy.append(int(data[3]))
				self.BPMdata_line.setData(self.BPMx,self.BPMy)
			if self.sw.secondsPassed() > 10:
				self.TempGraph.setXRange(self.sw.secondsPassed()-10,self.sw.secondsPassed())
				self.HumGraph.setXRange(self.sw.secondsPassed()-10,self.sw.secondsPassed())
				self.BPMGraph.setXRange(self.sw.secondsPassed()-10,self.sw.secondsPassed())
			if (float(data[0]) < 20 or float(data[0]) > 30) and self.ui.TempAlarm.isChecked():
				text = "Temperature is too low. Please check." if float(data[0]) < 20 else "Temperature is too high. Please check."
				self.ui.label_3.setText(text)
				if not pygame.mixer.get_busy():
					TempAlarm = pygame.mixer.Sound("assets\\BPMWarning.mp3")
					TempAlarm.play()
			elif (int(data[1]) < 20 or int(data[1]) > 80) and self.ui.HumAlarm.isChecked():
				text = "Humidity is too low. Please check." if int(data[1]) < 20 else "Humidity is too high. Please check."
				self.ui.label_3.setText(text)
				if not pygame.mixer.get_busy():
					HumAlarm = pygame.mixer.Sound("assets\\BPMWarning.mp3")
					HumAlarm.play()
			elif (int(data[3]) < 60 or int(data[3]) > 120) and self.ui.BPMAlarm.isChecked():
				text = "BPM is too low. Please check." if int(data[3]) < 60 else "BPM is too high. Please check."
				self.ui.label_3.setText(text)
				if not pygame.mixer.get_busy():
					BPMAlarm = pygame.mixer.Sound("assets\\BPMWarning.mp3")
					BPMAlarm.play()
			elif int(data[2]) < 50 and self.ui.Jaundice.isChecked():
				self.ui.label_3.setText("Check Blue Light.")
				if not pygame.mixer.get_busy():
					BlueLightAlarm = pygame.mixer.Sound("assets\\BPMWarning.mp3")
					BlueLightAlarm.play()
			else:
				self.ui.label_3.setText("")
				if pygame.mixer.get_busy():
					pygame.mixer.stop()
			users = pd.read_excel("assets\\Database\\users.xlsx", sheet_name=['Babies'])
			babies = users['Babies']
			for i,baby in enumerate(babies['username']):
				if baby == session.username:
					babiess = openpyxl.load_workbook('assets\\Database\\users.xlsx')
					babiesss = babiess["Babies"]
					if float(data[0]) < babies['minTemp'][i]:
						babiesss.cell(row=i+2,column=3).value = data[0]
						babiess.save(os.getcwd() + "\\assets\\Database\\users.xlsx")
						self.ui.label_18.setText(data[0])
					elif float(data[0]) > babies['maxTemp'][i]:
						babiesss.cell(row=i+2,column=4).value = data[0]
						babiess.save(os.getcwd() + "\\assets\\Database\\users.xlsx")
						self.ui.label_19.setText(data[0])
					if int(data[1]) < babies['minHum'][i]:
						babiesss.cell(row=i+2,column=5).value = data[1]
						babiess.save(os.getcwd() + "\\assets\\Database\\users.xlsx")
						self.ui.label_20.setText(data[1])
					elif int(data[1]) > babies['maxHum'][i]:
						babiesss.cell(row=i+2,column=6).value = data[1]
						babiess.save(os.getcwd() + "\\assets\\Database\\users.xlsx")
						self.ui.label_21.setText(data[1])
					if int(data[3]) < babies['minBPM'][i]:
						babiesss.cell(row=i+2,column=7).value = data[3]
						babiess.save(os.getcwd() + "\\assets\\Database\\users.xlsx")
						self.ui.label_23.setText(data[3])
					elif int(data[3]) > babies['maxBPM'][i]:
						babiesss.cell(row=i+2,column=8).value = data[3]
						babiess.save(os.getcwd() + "\\assets\\Database\\users.xlsx")
						self.ui.label_25.setText(data[3])
					break
			try:
				f = open("assets\\Database\\Logs\\" + session.username +'.txt','a')
			except:
				open("assets\\Database\\Logs\\" + session.username +'.txt','x')
				f = open("assets\\Database\\Logs\\" + session.username +'.txt','a')
			f.write('[' + str(datetime.now()) + ']:' + 'Temperature:' + data[0] + ', ' + 'Humidity:' + data[1] + ', ' + 'BPM:' + data[3] + '\n')
			f.close()
			

				

	def MonitoringMode(self):
		self.ui.stackedWidget_2.setCurrentIndex(0)

	def StatisticsMode(self):
		self.ui.stackedWidget_2.setCurrentIndex(1)

	def DetectJaundice(self):
		if self.ui.Jaundice.isChecked():
			self.serialPort.write('ON'.encode('utf-8'))
		else:
			self.serialPort.write('OFF'.encode('utf-8'))

	def LogOut(self):
		win = LoginWindow()
		win.setWindowTitle("Sign Up")
		win.setWindowIcon(QtGui.QIcon("assets\\865969.png"))
		win.show()
		self.serialPort.close()
		self.close()
		
	def closeEvent(self, a0: QtGui.QCloseEvent):
		self.serialPort.close()
		self.connected = False
		return super().closeEvent(a0)